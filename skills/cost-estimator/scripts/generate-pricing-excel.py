#!/usr/bin/env python3
"""
Generic JSON-to-styled-Excel converter for pricing reports.

Reads a JSON file describing the report structure and produces a formatted .xlsx.

Usage:
    python3 generate-pricing-excel.py --input report.json --output pricing.xlsx

Input JSON schema:
{
  "sheets": [
    {
      "name": "us-east-1",                              // Sheet tab name
      "columns": ["Module", "Service", ...],             // Header row
      "columnWidths": [20, 35, ...],                     // Column widths
      "rows": [                                          // Data rows
        {"values": ["Console", "ECS Fargate", ...], "module": "Console"},
        ...
      ],
      "totalRow": {"label": "Total", "value": 639.29},  // Optional total
      "assumptions": ["Deployed in us-east-1", ...]      // Optional Key Assumptions
    }
  ]
}

- "module" on each row → color-banding (same module = same color, auto-assigned).
- Numeric values → right-aligned with #,##0.00 format.
- "—" → centered dash (minimal cost).
- "totalRow" → bold dark summary row.
- "assumptions" → printed below the table as "Key Assumptions:" bullet list.

Requires: openpyxl
"""

import argparse
import json
import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    print("Error: openpyxl required. Install: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# --- Style constants ---

COLOR_PALETTE = [
    "E3F2FD", "E8F5E9", "FFF3E0", "F3E5F5",
    "FFEBEE", "E0F7FA", "FFF9C4", "F1F8E9",
]
HEADER_FILL = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
NUM_FMT = '#,##0.00'


def build_sheet(wb: Workbook, sheet_def: dict):
    name = sheet_def.get("name", "Sheet")[:31]
    ws = wb.create_sheet(title=name)

    columns = sheet_def.get("columns", [])
    col_widths = sheet_def.get("columnWidths", [])
    rows = sheet_def.get("rows", [])

    # Column widths
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # Header row
    for ci, header in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    # Module color assignment
    color_map = {}

    def module_fill(mod: str) -> PatternFill:
        if not mod:
            return PatternFill(fill_type=None)
        if mod not in color_map:
            color_map[mod] = COLOR_PALETTE[len(color_map) % len(COLOR_PALETTE)]
        c = color_map[mod]
        return PatternFill(start_color=c, end_color=c, fill_type="solid")

    # Data rows
    for ri, row_def in enumerate(rows, 2):
        values = row_def.get("values", [])
        mod = row_def.get("module", "")
        fill = module_fill(mod)

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            cell.fill = fill

            # Numeric formatting for price columns
            if isinstance(val, (int, float)):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            elif val == "—" or val == "—":
                cell.alignment = Alignment(horizontal="center")

        # Bold first column if it has a module name (first in group)
        if values and values[0]:
            ws.cell(row=ri, column=1).font = Font(bold=True)

    # Total row
    next_row = len(rows) + 2
    total_def = sheet_def.get("totalRow")
    if total_def:
        tr = next_row
        total_label = total_def.get("label", "Total")
        total_value = total_def.get("value", 0)

        for ci in range(1, len(columns) + 1):
            cell = ws.cell(row=tr, column=ci, value="")
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = BORDER

        ws.cell(row=tr, column=1, value=total_label).fill = TOTAL_FILL
        ws.cell(row=tr, column=1).font = TOTAL_FONT

        # Put total value in the price column (find "Price" or "Cost" in header)
        price_col = None
        for ci, header in enumerate(columns, 1):
            if "price" in header.lower() or "cost" in header.lower():
                price_col = ci
                break
        if price_col is None and len(columns) >= 5:
            price_col = 5

        if price_col:
            # Use SUM formula instead of hardcoded value for accuracy
            col_letter = chr(64 + price_col)
            data_start = 2
            data_end = len(rows) + 1
            sum_formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            tc = ws.cell(row=tr, column=price_col, value=sum_formula)
            tc.fill = TOTAL_FILL
            tc.font = TOTAL_FONT
            tc.number_format = NUM_FMT
            tc.alignment = Alignment(horizontal="right")

        next_row = tr + 2

    # Key Assumptions
    assumptions = sheet_def.get("assumptions", [])
    if assumptions:
        ar = next_row
        cell = ws.cell(row=ar, column=1, value="Key Assumptions:")
        cell.font = Font(bold=True, size=11)
        for i, text in enumerate(assumptions):
            ws.cell(row=ar + 1 + i, column=1, value=f"• {text}").font = Font(size=10)
            ws.merge_cells(start_row=ar + 1 + i, start_column=1,
                           end_row=ar + 1 + i, end_column=len(columns))


def main():
    parser = argparse.ArgumentParser(description="JSON → styled Excel pricing report")
    parser.add_argument("--input", required=True, help="Input JSON file describing the report")
    parser.add_argument("--output", required=True, help="Output .xlsx file path")
    args = parser.parse_args()

    with open(args.input) as f:
        report = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in report.get("sheets", []):
        build_sheet(wb, sheet_def)

    if not wb.sheetnames:
        print("Warning: no sheets defined in input JSON", file=sys.stderr)
        wb.create_sheet("Empty")

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"Written: {args.output} ({len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
